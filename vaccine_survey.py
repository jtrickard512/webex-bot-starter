import requests
import json
import time
from openpyxl import load_workbook
from requests.models import HTTPError
from requests_toolbelt.multipart.encoder import MultipartEncoder
from tqdm import tqdm

dest_filename = "C:/Users/josricka/Dev/webex-bot-starter/spreadsheets/Bot_Q3FY21_RD.xlsx"

wb = load_workbook(dest_filename)
ws = wb['Sheet1']

requests.packages.urllib3.disable_warnings()
session = requests.session()

f = open('config.json')
data = json.load(f)
c = open('vaccine.json')
card = json.load(c)


def send_realdeal_message(url,email,card):


    m = {'toPersonEmail':f'{email}',
            'text': 'test',
            'attachments': [card]}
    
    headers = {
        'Authorization':f'Bearer {bearer}',
    }

    response = session.post(url,json=m,headers=headers)
    response.raise_for_status()
    return response

bearer = data['token']

with tqdm(total=1) as pbar:
    for i,row in enumerate(ws.rows):
        values = [cell.value for cell in row]
        if i > 0 and values[3] != 'Sent':
            name = values[1]
            email = values[0]
            link = values[2]
            card['content']['body'][2]['text'] = f'Hello {name}'
            #card['content']['body'][4]['actions'][0]['url'] = link
            try:
                result = send_realdeal_message('https://webexapis.com/v1/messages',email,card)
                #print(i, result.status_code, time.time(), result.headers['Trackingid'])
                ws[f'D{i +1}'] = 'Sent'
                pbar.update(1)
            except HTTPError as e:
                error = True
                status_code = e.response.status_code
                while error:
                    if status_code == 429:
                        print('code', status_code)
                        print ('headers', e.response.headers)
                        print ('Sleeping for', e.response.headers['Retry-After'], 'seconds')
                        sleep_time = int(e.response.headers['Retry-After'])
                        while sleep_time > 10:
                            time.sleep(10)
                            sleep_time -= 10
                            print ('Asleep for', sleep_time, 'more seconds')
                        time.sleep(sleep_time)
                    elif status_code == 404:
                        ws[f'D{i +1}'] = 'Sent'
                        ws[f'E{i+1}'] = '404 error. Person not in Webex.'
                        wb.save(filename=dest_filename)
                        break
                    elif status_code == 500:
                        print ('500 error. Asleep for', 5, 'seconds')
                        time.sleep(5)
                    else:
                        wb.save(filename=dest_filename)
                        print(e, status_code)
                    print('Retrying message')
                    result = send_realdeal_message('https://webexapis.com/v1/messages',email,card)
                    print(i, result.status_code, time.time(), result.headers['Trackingid'])
                    if result.status_code == 200:
                        ws[f'D{i +1}'] = 'Sent'
                        pbar.update(1)
                        error = False
        else:
            pbar.update(1)
        
print('Spreadsheet saved')
wb.save(filename=dest_filename)




    