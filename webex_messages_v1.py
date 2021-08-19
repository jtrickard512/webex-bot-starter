import requests
import json
import time
from openpyxl import load_workbook
from requests.models import HTTPError
from requests_toolbelt.multipart.encoder import MultipartEncoder
from tqdm import tqdm

dest_filename = "C:/Users/josricka/Dev/webex-bot-starter/spreadsheets/Bot_Q3FY21_RD.xlsx"

wb = load_workbook(dest_filename)
ws = wb['Sheet1']


requests.packages.urllib3.disable_warnings()
session = requests.session()

f = open('config.json')
data = json.load(f)

def split_url_link(link):
    unique_id = link.split('/')[3]
    return unique_id

def invitation(name,link):
    id = split_url_link(link)
    link = f'https://listening.cisco.com/realdeal/{id}'
    
    message = f"""
    {name},
    \nYou're in luck, the survey is still open until end of day on Wednesday.  
    \nWe are also in luck because we still need your help.
    \nThis is your last chance to respond to the survey and let Cisco leaders know how you feel..
    \n[Access Survey]({link})
    \nThis nudge was also sent in email form, either link will take you to the same survey, choose your own adventure. Please check to make sure you are on VPN if you have any issues accessing the link. If you have any questions, you can ask for help by typing "help" in as a reply to this message." 
    """

    return message
    

def send_realdeal_message(url,email,message):

    m = MultipartEncoder({'toPersonEmail':f'{email}',
                                          'markdown': message })
    
    headers = {
        'Authorization':f'Bearer {bearer}',
        'Content-Type': m.content_type
    }

    response = session.post(url,data=m,headers=headers)
    response.raise_for_status()
    return response

bearer = data['token']

def new_func(dest_filename, wb, ws, invitation, send_realdeal_message):
    for i,row in tqdm(enumerate(ws.rows)):
        values = [cell.value for cell in row]
        if i > 0 and ws[f'F{i+1}'] != 'Sent':
            name = values[1]
            email = values[0]
            link = values[2]
            message = invitation(name,link)
            try:
                result = send_realdeal_message('https://webexapis.com/v1/messages',email,message)
                print(i, result.status_code, time.time(), result.headers['Trackingid'])
                ws[f'F{i +1}'] = 'Sent'
            except HTTPError as e:
                error = True
                status_code = e.response.status_code
                while error:
                    if status_code == 429:
                        print('code', status_code)
                        print ('headers', e.response.headers)
                        print ('Sleeping for', e.response.headers['Retry-After'], 'seconds')
                        sleep_time = int(e.response.headers['Retry-After'])
                        while sleep_time > 10:
                            time.sleep(10)
                            sleep_time -= 10
                            print ('Asleep for', sleep_time, 'more seconds')
                        time.sleep(sleep_time)
                    elif status_code == 404:
                        ws[f'F{i +1}'] = 'Sent'
                        ws[f'G{i+1}'] = '404 error. Person not in Webex.'
                        wb.save(filename=dest_filename)
                        break
                    elif status_code == 500:
                        print ('500 error. Asleep for', 5, 'seconds')
                        time.sleep(5)
                    else:
                        wb.save(filename=dest_filename)
                        print(e, status_code)
                    print('Retrying message')
                    result = send_realdeal_message('https://webexapis.com/v1/messages',email,message)
                    print(i, result.status_code, time.time(), result.headers['Trackingid'])
                    if result.status_code == 200:
                        ws[f'F{i +1}'] = 'Sent'
                        error = False


new_func(dest_filename, wb, ws, invitation, send_realdeal_message)
        
print('Spreadsheet saved')
wb.save(filename=dest_filename)




    